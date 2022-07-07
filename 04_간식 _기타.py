from bs4 import BeautifulSoup as bs
import requests as rq
from openpyxl import Workbook
import time , re , os , pyautogui


class Application:
    def __init__(self):
        # headers
        self.user_agent = 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/100.0.4896.57 Whale/3.14.133.23 Safari/537.36'
        self.headers = {'User-Agent': self.user_agent}


    def run(self)-> list:
        # 링크 주소 추출
        links = self.get_link()

        # 페이지 내 데이터 추출
        results = self.get_content(links=links)

        # 추출데이터 return 하기
        return results


    def get_link(self)-> list:
        url = 'https://www.fatsecret.kr/%EC%B9%BC%EB%A1%9C%EB%A6%AC-%EC%98%81%EC%96%91%EC%86%8C/%EA%B7%B8%EB%A3%B9/%EA%B0%84%EC%8B%9D'
        base_url = 'https://www.fatsecret.kr'

        response = rq.get(url, headers=self.headers)
        html = response.text
        soup = bs(html, 'html.parser')

        tag = soup.select('td.content a')

        lnks = []
        for x in tag:
            link = base_url + x.attrs['href']

            lnks.append(link)

        # 링크주소들 리턴
        return lnks

    def get_content(self,links:list)-> list:
        base_url = 'https://www.fatsecret.kr'

        save_data = []

        for link in links:
            response = rq.get(link, headers=self.headers)
            html = response.text
            soup = bs(html, 'html.parser')

            # 카테고리 이름 추출
            keword_title = soup.select('h1.title')[-1].text.strip()

            # 상세페이지 링크 태그 추출
            detail_links = soup.select('div.food_links > a')

            for detail_link in detail_links:
                link_detail = base_url + detail_link.attrs['href']

                response_detail = rq.get(link_detail, headers=self.headers)
                soup_detail = bs(response_detail.text, 'html.parser')

                # 음식명
                title = soup_detail.select_one('td.center h1').text.strip()

                # 열량 (Kcal)
                info_01 = int(soup_detail.select('div.factValue')[0].text.strip())

                # 지방
                info_02 = soup_detail.select('div.factValue')[1].text.strip()

                # 탄수화물
                info_03 = soup_detail.select('div.factValue')[2].text.strip()

                # 단백질
                info_04 = soup_detail.select('div.factValue')[-1].text.strip()

                # 열량 ( KJ )
                info_05 = soup_detail.select('div.nutrient.black.right.tRight')[0].text.strip()

                # 서빙사이즈
                serving_size = soup_detail.select_one('div.serving_size.black.serving_size_value').text.strip()

                # 칼로리 , 지방 , 탄수화물 , 단백질 태그 제외
                check = soup_detail.select('div.nutrient.right.tRight')
                check_list = []
                for x in check:
                    if 'black' not in str(x):
                        check_list.append(x)

                # 설탕당
                try:
                    if soup_detail.select('div.nutrient.left')[3].text == '설탕당':
                        info_06 = check_list[1].text.strip()
                    else:
                        info_06 = '-'
                except:
                    info_06 = '-'

                # 포화지방
                try:
                    if soup_detail.select('div.nutrient.left')[6].text == '포화지방':
                        info_07 = check_list[2].text.strip()
                    else:
                        info_07 = '-'
                except:
                    info_07 = '-'

                # 콜레스테롤
                try:
                    if soup_detail.select('div.nutrient.left')[9].text == '콜레스테롤':
                        info_8 = check_list[5].text.strip()
                    else:
                        info_8 = '-'
                except:
                    info_8 = '-'

                # 식이섬유
                try:
                    if soup_detail.select('div.nutrient.left')[10].text == '식이섬유':
                        info_9 = check_list[6].text.strip()
                    else:
                        info_9 = '-'
                except:
                    info_9 = '-'

                # 나트륨
                try:
                    if soup_detail.select('div.nutrient.left')[-2].text == '나트륨':
                        info_10 = check_list[7].text.strip()
                    else:
                        info_10 = '-'
                except:
                    info_10 = '-'

                # 칼슘
                try:
                    if soup_detail.select('div.nutrient.left')[-1].text == '칼슘':
                        info_11 = check_list[-1].text.strip()
                    else:
                        info_11 = '-'
                except:
                    info_11 = '-'

                # 데이터 저장
                save_data.append([
                    keword_title,
                    title,
                    serving_size,
                    info_05,
                    info_01,
                    info_02,
                    info_03,
                    info_04,
                    info_06,
                    info_07,
                    info_8,
                    info_9,
                    info_10,
                    info_11
                ])
                print(f"""
    카테고리 : {keword_title}
    음식명 : {title}
    서빙사이즈 : {serving_size}
    열량 (KJ) : {info_05}
    열량 (Kcal) : {info_01}
    지방 : {info_02}
    탄수화물 : {info_03}
    단백질 : {info_04}
    설탕당 : {info_06}
    포화지방 : {info_07}
    콜레스테롤 : {info_8}
    식이섬유 : {info_9}
    나트륨 : {info_10}
    칼슘 : {info_11}
    """)

        # 데이터 리턴
        return save_data


class OpenPyXL:
    def __init__(self):
        # WorkBook
        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws.append([
            '카테고리',
            '음식명',
            '서빙사이즈',
            '열량 (KJ)',
            '열량 (Kcal)',
            '지방',
            '탄수화물',
            '단백질',
            '설탕당',
            '포화지방',
            '콜레스테롤',
            '식이섬유',
            '나트륨',
            '칼슘'
        ])

        # Application class의 run 함수 실행 후, 리턴 값 멤버변수로 정의
        self.results = Application().run()

        # 파일저장 메소드 실행
        self.savefile()


    def savefile(self)-> None:
        row = 2
        for x in self.results:
            self.ws[f"A{row}"] = x[0]
            self.ws[f"B{row}"] = x[1]
            self.ws[f"C{row}"] = x[2]
            self.ws[f"D{row}"] = x[3]
            self.ws[f"E{row}"] = x[4]
            self.ws[f"F{row}"] = x[5]
            self.ws[f"G{row}"] = x[6]
            self.ws[f"H{row}"] = x[7]
            self.ws[f"I{row}"] = x[8]
            self.ws[f"J{row}"] = x[9]
            self.ws[f"K{row}"] = x[10]
            self.ws[f"L{row}"] = x[11]
            self.ws[f"M{row}"] = x[12]
            self.ws[f"N{row}"] = x[-1]

            row += 1

        # 저장경로 지정
        savePath = os.path.abspath('fatsecret')
        fileName = '간식_기타.xlsx'

        if not os.path.exists(savePath):
            os.mkdir(savePath)
        self.wb.save(os.path.join(savePath, fileName))
        self.wb.close()

        pyautogui.alert(f'파일추출완료\n\n{savePath}')






if __name__ == '__main__' :
    app = OpenPyXL()



