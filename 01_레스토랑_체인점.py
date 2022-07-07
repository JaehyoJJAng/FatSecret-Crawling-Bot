import pyautogui
from bs4 import BeautifulSoup as bs
import requests as rq
from openpyxl import Workbook
import time ,re ,os
import urllib.parse as rep
from conf import get_word_list

class Application:
    def __init__(self):
        self.word_list : list = list(get_word_list('word_list'))

        # headers
        self.headers = {
            'User-Agent':'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/100.0.4896.57 Whale/3.14.133.23 Safari/537.36'
        }

    def run(self):
        # 브랜드별 링크 추출
        links = self.get_food_link()

        # 브랜드별 상세페이지 링크 추출
        detail_links = self.get_detail_link(links=links)

        # 상세페이지 내 데이터 추출
        results = self.get_content(detail_links=detail_links)

        # 추출데이터 return
        return results

    def get_food_link(self)-> list:
        base_url = 'https://www.fatsecret.kr'
        food_link = []

        for word in self.word_list:
            print(f"{'=' * 50} {word} {'=' * 50}")
            quote_word = rep.quote_plus(word)
            URL = f'https://www.fatsecret.kr/Default.aspx?pa=brands&f={word}&t=2'

            response = rq.get(URL, headers=self.headers)
            html = response.text
            soup = bs(html, 'html.parser')

            a_tags : list = soup.select('h2 > a')

            for a in a_tags:
                food_link.append(base_url + a.attrs['href'])

        return food_link

    def get_detail_link(self,links)-> list:
        detail_link_list = []
        base_url = 'https://www.fatsecret.kr'

        for link in links  :
            response = rq.get(link,headers=self.headers)
            html = response.text
            soup = bs(html,'html.parser')

            title = soup.select_one('td.center h1').text.strip()

            detail_lnks = soup.select('a.prominent')

            for detail_link in detail_lnks :
                detail_link_list.append(base_url + detail_link.attrs['href'])

        return detail_link_list

    def get_content(self,detail_links)-> list:
        save_data = []

        for link in detail_links:
            response = rq.get(link, headers=self.headers)
            html = response.text
            soup = bs(html, 'html.parser')

            # 브랜드명
            brand_name = soup.select_one('h2.manufacturer > a').text.strip()

            # 메뉴명
            menu_name = soup.select_one('td.center h1').text.strip()

            # 열량 (Kcal)
            info_01 = int(soup.select('div.factValue')[0].text.strip())

            # 열량 (KJ)
            info_05 = soup.select('div.nutrient.black.right.tRight')[0].text.strip()

            # 지방
            info_02 = soup.select('div.factValue')[1].text.strip()

            # 탄수화물
            info_03 = soup.select('div.factValue')[2].text.strip()

            # 단백질
            info_04 = soup.select('div.factValue')[-1].text.strip()

            # 서빙사이즈
            serving_size = soup.select_one('div.serving_size.black.serving_size_value').text.strip()

            # 탄수화물
            info_06 = soup.select('div.nutrient.black.right.tRight')[1].text.strip()

            # 단백질
            info_07 = soup.select('div.nutrient.black.right.tRight')[2].text.strip()

            # 지방
            info_08 = soup.select('div.nutrient.black.right.tRight')[-1].text.strip()

            check = soup.select('div.nutrient.right.tRight')
            check_list = []
            for x in check:
                if 'black' not in str(x):
                    check_list.append(x)

            # 설탕당
            try:
                if soup.select('div.nutrient.left')[3].text == '설탕당':
                    info_06_sweety = check_list[1].text.strip()
                else:
                    info_06_sweety = '-'
            except:
                info_06_sweety = '-'

            # 포화지방
            try:
                if soup.select('div.nutrient.left')[6].text == '포화지방':
                    info_08_detail = check_list[2].text.strip()
                else:
                    info_08_detail = '-'
            except:
                info_08_detail = '-'

            try:
                # 나트륨
                if soup.select('div.nutrient.left')[-1].text == '나트륨':
                    info_09 = check_list[-1].text.strip()
                else:
                    info_09 = '-'
            except:
                info_09 = '-'

            save_data.append(
                [brand_name, menu_name, info_05, info_01, info_02, info_03, info_04, serving_size, info_06_sweety,
                 info_08_detail, info_09])
            print(f"""
    브랜드 명 : {brand_name}
    메뉴명 : {menu_name}
    열량(KJ) : {info_05}
    열량(Kcal) : {info_01}
    지방 : {info_02}
    탄수화물 : {info_03}
    단백질 : {info_04}
    서빙사이즈 : {serving_size}
    설탕당 : {info_06_sweety}
    포화지방 : {info_08_detail}
    나트륨 : {info_09}
    """)
        return save_data


class OpenPyXL:
    def __init__(self):
        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws.title = '음식정보'
        self.ws.append([
            '브랜드명','메뉴명','열량(KJ)','열량(Kcal)','지방','탄수화물','단백질','서빙사이즈','설탕당','포화지방','나트륨'
        ])


        # 추출데이터
        self.results = Application().run()

        # 파일저장 메소드 실행
        self.savefile()

    def savefile(self)-> None:
        row = 2
        for x in self.results:
            self.ws[f"A{row}"] = x[0]
            self.ws[f"B{row}"] = x[1]
            self.ws[f"C{row}"] = x[2]
            self.ws[f"D{row}"] = x[3]
            self.ws[f"E{row}"] = x[4]
            self.ws[f"F{row}"] = x[5]
            self.ws[f"G{row}"] = x[6]
            self.ws[f"H{row}"] = x[7]
            self.ws[f"I{row}"] = x[8]
            self.ws[f"J{row}"] = x[9]
            self.ws[f"K{row}"] = x[10]

            row += 1

        savePath = os.path.abspath('fatsecret')
        fileName = '레스토랑&체인.xlsx'
        if not os.path.exists(savePath):
            os.mkdir(savePath)

        self.wb.save(os.path.join(savePath, fileName))
        self.wb.close()

        pyautogui.alert(f'파일저장 완료!\n\n{savePath}')








if __name__ == '__main__' :
    app = OpenPyXL()

